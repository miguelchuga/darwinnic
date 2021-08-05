# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl import Workbook

import base64

from odoo import models, fields, api

class GenerateStockWizard(models.Model):
    
    _name = "rep_planilla.generate_planilla"
    
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')
    por_tipo_planilla = fields.Boolean(string='Agrupar por Tipo de Planilla', default=False, required=False)
    planillas_id = fields.Many2one('hr.payslip.run', string='Planilla')
    categoria_employee_id = fields.Many2one('hr.employee.category', string='Tipo de Planilla', required=False)

    def generate_file(self):

        _planilla_id = self.planillas_id._ids[0]
        _tipo_planilla = self.categoria_employee_id._ids[0]
                    
        this = self.id
         
        fileobj = NamedTemporaryFile('w+b')
        
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))

        user = self.env.user
         
        wb = Workbook()

        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        
#--------------------------------------------------------#
# ---------- DETALLE DE LA PLANILLA ---------------------#
#--------------------------------------------------------#

        ws.title = "detalle"

        ws['A1'].value = "Departamento"
        ws['B1'].value = "Puesto"
        ws['C1'].value = "Nombre"
        ws['D1'].value = "Concepto" 
        ws['E1'].value = "Total"
     
        al1 = Alignment(horizontal="left", vertical="center")
        al2 = Alignment(horizontal="right", vertical="center")
        font = Font(size=12, bold=True, italic=True, color="17202A") #, color="4777AD" color="E0D7D7" color="E0D7D7"
        fill = PatternFill("solid", fgColor="CACFD2")   #CACFD2
        thin = Side(border_style="thin", color="CACFD2")    #color="000000" gColor="C499C3"
        double = Side(border_style="double", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        self.style_range(ws, 'A1:E1', border=border, alignment=al1, font=font, fill=fill)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 20
        
        sql = """
             SELECT hpr.id,hp.id,hp.struct_id,hp.employee_id,hp.contract_id,hp.company_id,rc.name empresa, --Lines de detalle Nominas
                 hd.name departamento,hj.name puesto ,he.name_related empleado, hc.date_start,-- Datos de empleados
                 hpl.sequence,hpl.name, hpl.total-- Datos de los calculos de la nomina
               FROM hr_payslip_run hpr 
               JOIN hr_payslip hp ON hp.payslip_run_id = hpr.id
               JOIN hr_contract hc ON hc.id = hp.contract_id
               JOIN res_company rc ON rc.id = hp.company_id
               JOIN hr_employee he ON hp.employee_id = he.id 
               LEFT JOIN hr_department hd ON he.department_id = hd.id
               LEFT JOIN hr_job hj ON hj.id = he.job_id
               JOIN hr_payslip_line hpl ON hp.id = hpl.slip_id
              WHERE hpr.id = %s AND hpl.total > 0
             ORDER BY he.name_related, hpl.sequence
            """

        self.env.cr.execute(sql,(_planilla_id,))
         
        font = Font(size=11, bold=False, italic=False)

        row = 2
        
        for query_line in self.env.cr.dictfetchall():

            ws.cell(row=row, column=1).value  = query_line['departamento'] or None;
            ws.cell(row=row, column=2).value  = query_line['puesto'] or None;
            ws.cell(row=row, column=3).value  = query_line['empleado'] or None;
            ws.cell(row=row, column=4).value  = query_line['name'] or None;
            
            celda = ws.cell(row=row, column=1).coordinate+':'+ws.cell(row=row, column=4).coordinate
            self.style_range(ws, celda, alignment=al1, font=font)
            
            ws.cell(row=row, column=5).value  = round(query_line['total'],2) or 0;
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            
            celda = ws.cell(row=row, column=5).coordinate+':'+ws.cell(row=row, column=5).coordinate
            self.style_range(ws, celda, alignment=al2, font=font)
            
            row += 1

#--------------------------------------------------------#
# ---------- REPORTE DE DEPOSITOS          --------------#
#--------------------------------------------------------#

        # Cuentas Bancarias para depositos.
        ws = wb.create_sheet()
        
        ws.title = "sincronizacion"

        ws['A1'].value = "Codigo"
        ws['B1'].value = "Nombre"
        ws['C1'].value = "Cuenta"
        ws['D1'].value = "Empleado"
        ws['E1'].value = "Total" 

        al1 = Alignment(horizontal="left", vertical="center")
        al2 = Alignment(horizontal="right", vertical="center")
        font = Font(size=12, bold=True, italic=True, color="17202A") #, color="4777AD" color="E0D7D7" color="E0D7D7"
        fill = PatternFill("solid", fgColor="CACFD2")   #CACFD2
        thin = Side(border_style="thin", color="CACFD2")    #color="000000" gColor="C499C3"
        double = Side(border_style="double", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        self.style_range(ws, 'A1:E1', border=border, alignment=al1, font=font, fill=fill)
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 15
        
        sql = """
            SELECT hpr.id, fp.code as codigo, fp.name as nombre, he.cuenta_bancaria,he.name_related , -- Datos de empleados
                   hpl.total -- Datos de los calculos de la nomina
             FROM hr_payslip_run hpr 
                  JOIN hr_payslip hp ON hp.payslip_run_id = hpr.id
                  JOIN res_company rc ON rc.id = hp.company_id
                  JOIN hr_employee he ON hp.employee_id = he.id
                  LEFT JOIN  forma_pago fp ON fp.id = he.forma_pago_id
                  LEFT JOIN hr_department hd ON he.department_id = hd.id
                  LEFT JOIN hr_job hj ON hj.id = he.job_id
                  JOIN hr_payslip_line hpl ON hp.id = hpl.slip_id
                  JOIN hr_salary_rule hsr ON hsr.id = hpl.salary_rule_id and hsr.imprime_liquido is True
            WHERE fp.code IN ('1', '2') AND hpl.total > 0 AND hpr.id = %s
            ORDER BY fp.name, he.name_related
            """
            
        self.env.cr.execute(sql,(_planilla_id,))
        
        row = 2
        
        font = Font(size=11, bold=False, italic=False)
        
        for query_line in self.env.cr.dictfetchall():
            
            ws.cell(row=row, column=1).value  = query_line['codigo'] or None;     
            ws.cell(row=row, column=2).value  = query_line['nombre'] or None;           
            ws.cell(row=row, column=3).value  = query_line['cuenta_bancaria'] or None;
            ws.cell(row=row, column=4).value  = query_line['name_related'] or None;
            
            celda = ws.cell(row=row, column=1).coordinate+':'+ws.cell(row=row, column=4).coordinate
            self.style_range(ws, celda, alignment=al1, font=font)
            
            ws.cell(row=row, column=5).value  = round(query_line['total'],2) or 0;
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            
            celda = ws.cell(row=row, column=5).coordinate+':'+ws.cell(row=row, column=5).coordinate
            self.style_range(ws, celda, alignment=al2, font=font)
            
            row += 1
                    
#--------------------------------------------------------#
# ---------- IMPORTAR CHEQUES  --------------------------#
#--------------------------------------------------------#

        ws = wb.create_sheet()
        
        ws.title = "cheques"
        
        ws['A1'].value = "partner_id"
        ws['B1'].value = "amount"
        ws['C1'].value = "payment_date" 
        ws['D1'].value = "journal_id" 
        ws['E1'].value = "communication" 
        ws['F1'].value = "x_account_id" 
        ws['G1'].value = "payment_type"
        ws['H1'].value = "Payment Method Type"  
 
        al1 = Alignment(horizontal="left", vertical="center")
        al2 = Alignment(horizontal="right", vertical="center")
        font = Font(size=12, bold=True, italic=True, color="17202A") #, color="4777AD" color="E0D7D7" color="E0D7D7"
        fill = PatternFill("solid", fgColor="CACFD2")   #CACFD2
        thin = Side(border_style="thin", color="CACFD2")    #color="000000" gColor="C499C3"
        double = Side(border_style="double", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        self.style_range(ws, 'A1:H1', border=border, alignment=al1, font=font, fill=fill)
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 24
        
        sql = """
            SELECT rc.name company_id,he.name_related partner_id, -- Datos de empleados
                   hpl.total amount, hpr.date_end Date, aj.name Journal, aa.code account_id,
                   hpr.name narration, aas.code voucher_lines_account,
                   aaa.code voucher_lines_analytic_account,
                   hpl.total voucher_lines_amount,
                   'Enviar dinero' payment_type,
                   'Cheque' payment_method_id
             FROM hr_payslip_run hpr 
                  JOIN hr_payslip hp ON hp.payslip_run_id = hpr.id
                  JOIN res_company rc ON rc.id = hp.company_id
                  JOIN hr_employee he ON hp.employee_id = he.id
                  JOIN hr_contract hc ON hc.employee_id = he.id
                  JOIN account_analytic_account aaa ON aaa.id = hc.analytic_account_id                  
                  JOIN account_journal aj ON aj.id = hc.diario_cheques_id
                  JOIN account_account aa ON aa.id = aj.default_credit_account_id
                  JOIN forma_pago fp ON fp.id = he.forma_pago_id
                  JOIN hr_payslip_line hpl ON hp.id = hpl.slip_id
                  JOIN hr_salary_rule hsr ON hsr.id = hpl.salary_rule_id and hsr.imprime_liquido is True
                  JOIN account_account aas ON aas.id = hsr.account_credit
            WHERE fp.code IN ('3') AND hpl.total > 0 AND hpr.id = %s
            ORDER BY fp.name, he.name_related
            """

        self.env.cr.execute(sql,(_planilla_id,))
        
        row = 2
        
        font = Font(size=11, bold=False, italic=False)
         
        for query_line in self.env.cr.dictfetchall():    

            ws.cell(row=row, column=1).value  = query_line['partner_id'] or None;
            
            celda = ws.cell(row=row, column=1).coordinate+':'+ws.cell(row=row, column=2).coordinate
            self.style_range(ws, celda, alignment=al1, font=font)
            
            ws.cell(row=row, column=2).value  = query_line['amount'] or None;
            ws.cell(row=row, column=2).number_format = '#,##0.00'

            celda = ws.cell(row=row, column=2).coordinate+':'+ws.cell(row=row, column=2).coordinate
            self.style_range(ws, celda, alignment=al2, font=font)
            
            ws.cell(row=row, column=3).value  = None;
            ws.cell(row=row, column=4).value  = query_line['journal'] or None;
            ws.cell(row=row, column=5).value  = query_line['narration'] or None;
            ws.cell(row=row, column=6).value  = query_line['voucher_lines_account'] or None;
            
            celda = ws.cell(row=row, column=3).coordinate+':'+ws.cell(row=row, column=6).coordinate
            self.style_range(ws, celda, alignment=al1, font=font)
            
            ws.cell(row=row, column=7).value  = query_line['payment_type'] or None;            
            ws.cell(row=row, column=8).value  = query_line['payment_method_id'] or None;
            
            row += 1
            
#--------------------------------------------------------#
# ---------- Nota de débito del Banco -------------------#
#--------------------------------------------------------#

        ws = wb.create_sheet()
        
        ws.title = "debito"

        ws['A1'].value = "Date"
        ws['B1'].value = "Journal"
        ws['C1'].value = "ref"
        ws['D1'].value = "Journal Items / Name" 
        ws['E1'].value = "Journal Items / Analytic Account"
        ws['F1'].value = "Journal Items / Partner"
        ws['G1'].value = "Journal Items / Account"
        ws['H1'].value = "Journal Items / Debit"
        ws['I1'].value = "Journal Items / Credit"

        al1 = Alignment(horizontal="left", vertical="center")
        al2 = Alignment(horizontal="right", vertical="center")
        font = Font(size=12, bold=True, italic=True, color="17202A") #, color="4777AD" color="E0D7D7" color="E0D7D7"
        fill = PatternFill("solid", fgColor="CACFD2")   #CACFD2
        thin = Side(border_style="thin", color="CACFD2")    #color="000000" gColor="C499C3"
        double = Side(border_style="double", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        self.style_range(ws, 'A1:I1', border=border, alignment=al1, font=font, fill=fill)
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 24
        
        # Obtiene Cargos de Pardida de Planilla.
        sql = """
            SELECT hpr.id, fp.code as codigo, fp.name as nombre, he.cuenta_bancaria, he.name_related as nombre_empleado, -- Datos de empleados
                   hp.date_from, hpl.total as cargo, -- Datos de los calculos de la nomina
                   aaa.code analytic_account
             FROM hr_payslip_run hpr 
                  JOIN hr_payslip hp ON hp.payslip_run_id = hpr.id
                  JOIN res_company rc ON rc.id = hp.company_id
                  JOIN hr_employee he ON hp.employee_id = he.id
                  LEFT JOIN forma_pago fp ON fp.id = he.forma_pago_id
                  LEFT JOIN hr_department hd ON he.department_id = hd.id
                  LEFT JOIN hr_job hj ON hj.id = he.job_id
                  JOIN hr_payslip_line hpl ON hp.id = hpl.slip_id
                  JOIN hr_salary_rule hsr ON hsr.id = hpl.salary_rule_id and hsr.imprime_liquido is True
                  JOIN hr_contract hc ON hc.employee_id = he.id
                  JOIN account_analytic_account aaa ON aaa.id = hc.analytic_account_id
            WHERE fp.code IN ('1', '2') AND hpl.total > 0 AND hpr.id = %s
            ORDER BY fp.name, he.name_related         
            """
            
        self.env.cr.execute(sql,(_planilla_id,))
        
        al1 = Alignment(horizontal="left", vertical="center")
        font = Font(size=11, bold=False, italic=False) #, color="4777AD"

        row = 2        
      
        wtotal = 0
        
        for query_line in self.env.cr.dictfetchall():

            if (row == 2):

                fs = query_line['date_from'].split('-')
                _fecha=((fs[2])+"/"+fs[1]+"/"+fs[0])
                            
                ws.cell(row=row, column=1).value = _fecha or None;
                ws.cell(row=row, column=2).value = self.env['hr.payslip.run'].browse(_planilla_id).journal_id.name
                ws.cell(row=row, column=3).value = self.env['hr.payslip.run'].browse(_planilla_id).name    
          
            ws.cell(row=row, column=4).value  = self.env['hr.payslip.run'].browse(_planilla_id).journal_id.name;
            ws.cell(row=row, column=5).value  = query_line['analytic_account'] or None;
            ws.cell(row=row, column=6).value  = query_line['nombre_empleado'] or None;
            
            # Cuenta Sueldos por Pagar, está asignada a la regla salarial Liquido a Recibir,            
            ws.cell(row=row, column=7).value  = '33010405';            
            ws.cell(row=row, column=8).value  = round(query_line['cargo'],2) or 0;
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=9).value  = None;
            ws.cell(row=row, column=9).number_format = '#,##0.00'
            
            celda = ws.cell(row=row, column=4).coordinate+':'+ws.cell(row=row, column=4).coordinate
            self.style_range(ws, celda, alignment=al1, font=font)
            
            row += 1
            
            wtotal = wtotal + query_line['cargo']
            
        if (wtotal != 0):
            
            # Cuenta del banco usado para pagar.
            ws.cell(row=row, column=4).value = self.env['hr.payslip.run'].browse(_planilla_id).journal_id.name
            ws.cell(row=row, column=7).value  = '14020106'; 
            ws.cell(row=row, column=8).value  = None;
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=9).value  = round(wtotal,2) or 0;
            ws.cell(row=row, column=9).number_format = '#,##0.00'
            
#--------------------------------------------------------#
# ---------- ----------PLANILLA---- ---------------------#
#--------------------------------------------------------#

        # Reporte de Planilla.        
        ws = wb.create_sheet()
        ws.title = "Planilla"

        al1 = Alignment(horizontal="left", vertical="center")
        al2 = Alignment(horizontal="right", vertical="center")
        font = Font(size=12, bold=True, italic=True, color="17202A") #, color="4777AD" color="E0D7D7" color="E0D7D7"
        fill = PatternFill("solid", fgColor="CACFD2")   #CACFD2
        thin = Side(border_style="thin", color="CACFD2")    #color="000000" gColor="C499C3"
        double = Side(border_style="double", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        self.style_range(ws, 'A1:D1', border=border, alignment=al1, font=font, fill=fill)
        
        # Lee los titulos de las columnas del reporte en base a la configuración de las reglas salariales.
        sql_titulos = """
            SELECT t.id,hsr.orden_columna,hsr.titulos_columnas
              FROM (SELECT hpr.id id,hpl.salary_rule_id regla
                      FROM hr_payslip_run hpr 
                      JOIN hr_payslip hp ON hp.payslip_run_id = hpr.id
                      JOIN hr_payslip_line hpl ON hp.id = hpl.slip_id
                    GROUP BY hpr.id, hpl.salary_rule_id) t
              JOIN hr_salary_rule hsr ON hsr.id = t.regla
            WHERE hsr.imprime_reporte IS True AND t.id = %s 
            ORDER BY hsr.orden_columna
            """

        self.env.cr.execute(sql_titulos,(_planilla_id,))
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8

        ws.cell(row=5, column=1).value  = "Puesto" or None
        ws.cell(row=5, column=2).value  = "Nombre" or None
        ws.cell(row=5, column=3).value  = "Fecha Ingreso" or None
        ws.cell(row=5, column=4).value  = "Total" or 0
        
        # Registra lista de letras de columnas.
        sub_cell=[]
        
        # Registra totales de grupo.
        total_cell=[]
        
        reg = 3
        end_cell = 'A1:A1' ## tiene este rango para que no de error cuando no trae titulos.
        
        col_suma = 'E'
        
        total_cell.append('')
        
        # Leer cada regla salarial para colocar el título de la columna.
        for query_line in self.env.cr.dictfetchall():
            
            # Obtiene la columna final.
            end_cell = ws.cell(row=5, column=reg+1)
        
            # Título de la regla salarial.
            ws.cell(row=5, column=reg+1).value  = query_line['titulos_columnas'] or None                        
                                          
            # Formato para títulos.      
            rango = ws.cell(row=5, column=reg+1).coordinate+':'+ws.cell(row=5, column=reg+1).coordinate            
            self.style_range(ws, rango, border=border, alignment=al1, font=font, fill=fill)
            
            # Si hay título en la regla salarial.
            if query_line['titulos_columnas']:
            
                ws.column_dimensions[ws.cell(row=5, column=reg+1).column].width = len(query_line['titulos_columnas'])+6
                
            # Si no hay título en la regla salarial.
            else:
                
                ws.column_dimensions[ws.cell(row=5, column=reg+1).column].width = 5
            
            # Agrega todos los títulos de columna en la fila 5. 
            sub_cell.append(ws.cell(row=5, column=reg+1).column)         
            
            # Inicializa todos los totales de columna.
            total_cell.append('')
              
            reg += 1

        rango = 'A5:D5'
        last_col_sup = ws.dimensions[3]
        self.style_range(ws, rango, border=border, alignment=al1, font=font, fill=fill)

        # Nombre de la empresa.
        al1 = Alignment(horizontal="center", vertical="center")
        font = Font(size=20, bold=True, italic=True, color="17202A") 
        
        ws['A1'].value = self.env.user.company_id.name 
        
        self.style_range(ws, 'A1:D1', alignment=al1, font=font)
        
        ws.row_dimensions[1].height = 25

        # Nombre de la planilla.        
        ws['A2'].value = self.env['hr.payslip.run'].browse(_planilla_id).name
        
        self.style_range(ws, 'A2:D2', alignment=al1, font=font)
         
        ws.row_dimensions[2].height = 25
         
        # Si se quiere agrupar por tipo de planilla.
        if self.por_tipo_planilla:
            
            # Detalle de reglas salariales de esta nómina.
            sql_detalle = """
            
                SELECT d.planilla, d.empresa, d.categoria_planilla, d.departamento, d.puesto, d.empleado, d.date_start, d.total, d.nregla 
                  FROM (SELECT hpr.id planilla,rc.name empresa,
                               hd.name departamento,hj.name puesto, he.name_related empleado, hc.date_start, ecr.category_id, hec.name AS categoria_planilla, 
                               hpl.name nombre_regla, hpl.total,-- Datos de los calculos de la nomina
                               hsr.orden_columna orden, hsr.name nregla --orde de la columnas
                    FROM hr_payslip_run hpr 
                    JOIN hr_payslip hp ON hp.payslip_run_id = hpr.id
                    JOIN hr_contract hc ON hc.id = hp.contract_id
                    JOIN res_company rc ON rc.id = hp.company_id
                    JOIN hr_employee he ON hp.employee_id = he.id 
                    LEFT JOIN hr_department hd ON he.department_id = hd.id
                    LEFT JOIN hr_job hj ON hj.id = he.job_id
                    JOIN hr_payslip_line hpl ON hp.id = hpl.slip_id
                    JOIN hr_salary_rule hsr ON hsr.id = hpl.salary_rule_id and hsr.imprime_reporte is True
                    JOIN employee_category_rel ecr ON he.id = ecr.emp_id
                    JOIN hr_employee_category hec ON ecr.category_id = hec.id
                   WHERE hpr.id = %s
                ORDER BY hec.name, hd.name, he.name_related, hsr.orden_columna) d
            """
    
            self.env.cr.execute(sql_detalle,(_planilla_id,))
            
        # Si no se quiere agrupar por tipo de planilla.
        else:
                    
            # Si se quiere sólo este tipo de planilla.
            if  _tipo_planilla:
                
                # Detalle de reglas salariales de esta nómina.
                sql_detalle = """
                    SELECT d.planilla, d.empresa, d.categoria_planilla, d.departamento, d.puesto, d.empleado, d.date_start, d.total, d.nregla, d.category_id
                      FROM (SELECT hpr.id planilla, rc.name empresa , hec.name AS categoria_planilla
                                 , hd.name departamento, hj.name puesto, he.name_related empleado, hc.date_start, ecr.category_id 
                                 , hpl.name nombre_regla, hpl.total -- Datos de los calculos de la nomina
                                 , hsr.orden_columna orden, hsr.name nregla --orde de la columnas
                        FROM hr_payslip_run hpr 
                        JOIN hr_payslip hp ON hp.payslip_run_id = hpr.id
                        JOIN hr_contract hc ON hc.id = hp.contract_id
                        JOIN res_company rc ON rc.id = hp.company_id
                        JOIN hr_employee he ON hp.employee_id = he.id 
                        LEFT JOIN hr_department hd ON he.department_id = hd.id
                        LEFT JOIN hr_job hj ON hj.id = he.job_id
                        JOIN hr_payslip_line hpl ON hp.id = hpl.slip_id
                        JOIN hr_salary_rule hsr ON hsr.id = hpl.salary_rule_id and hsr.imprime_reporte is True
                        JOIN employee_category_rel ecr ON he.id = ecr.emp_id
                        JOIN hr_employee_category hec ON ecr.category_id = hec.id
                       WHERE hpr.id = %s AND hec.id = %s
                    ORDER BY hec.name, hd.name, he.name_related, hsr.orden_columna) d
                """
                
                self.env.cr.execute(sql_detalle,(_planilla_id, _tipo_planilla,))
                
            # Si no se inidica tipo de planilla.
            else:
                
                # Detalle de reglas salariales de esta nómina.
                sql_detalle = """
                    SELECT d.planilla, d.empresa, d.categoria_planilla, d.departamento, d.puesto, d.empleado, d.date_start, d.total, d.nregla, d.category_id
                      FROM (SELECT hpr.id planilla, rc.name empresa , hec.name AS categoria_planilla
                                 , hd.name departamento, hj.name puesto, he.name_related empleado, hc.date_start, ecr.category_id 
                                 , hpl.name nombre_regla, hpl.total -- Datos de los calculos de la nomina
                                 , hsr.orden_columna orden, hsr.name nregla --orde de la columnas
                        FROM hr_payslip_run hpr 
                        JOIN hr_payslip hp ON hp.payslip_run_id = hpr.id
                        JOIN hr_contract hc ON hc.id = hp.contract_id
                        JOIN res_company rc ON rc.id = hp.company_id
                        JOIN hr_employee he ON hp.employee_id = he.id 
                        LEFT JOIN hr_department hd ON he.department_id = hd.id
                        LEFT JOIN hr_job hj ON hj.id = he.job_id
                        JOIN hr_payslip_line hpl ON hp.id = hpl.slip_id
                        JOIN hr_salary_rule hsr ON hsr.id = hpl.salary_rule_id and hsr.imprime_reporte is True
                        JOIN employee_category_rel ecr ON he.id = ecr.emp_id
                        JOIN hr_employee_category hec ON ecr.category_id = hec.id
                       WHERE hpr.id = %s
                    ORDER BY hec.name, hd.name, he.name_related, hsr.orden_columna) d
                """
    
                self.env.cr.execute(sql_detalle,(_planilla_id,))
        
        # Columna en la que inicia a imprimir las reglas salariales.
        col = 3
        # Fila en la que están los títulos de las reglas salariales.
        row = 5
        # Para saber que es el primer empleado.
        _empleado = ''
        # Para saber que es el primer departamento.
        _departamento = ''
        # Para saber que es el primer tipo de planilla.
        _tipo_planilla = ''
        
        # Configuración de los títulos.
        al1 = Alignment(horizontal="left", vertical="center")
        al2 = Alignment(horizontal="right", vertical="center")
        font = Font(size=10, bold=True, italic=True, color="17202A") 
        
        # Configuración del detalle.
        al3 = Alignment(horizontal="left", vertical="center")
        al4 = Alignment(horizontal="right", vertical="center")
        font2 = Font(size=10, bold=False, italic=False, color="190707") 
        
        grupo_suma = []
        _primer_categoria = True
        _segundo_categoria = True
        _primer_departamento = True
        _segundo_departamento = True
        _cambia_categoria = False

        # Si se debe agrupar por tipo de planilla.
        if self.por_tipo_planilla:
            
            # Inicia a contar el primer regitro de detalle.
            first_row_sup = 10
                    
            # Lee cada registro de nómina.
            for query_line in self.env.cr.dictfetchall():
                                            
                # Si el empleado es diferente.
                if _empleado != query_line['empleado']:
                    
                    # En columna 4 inicia a imprimir reglas salariales del empleado.
                    col = 4
                    
                    # Avanza una línea.
                    row += 1
                    
                # Si el empleado es igual.
                else:
                    
                    # Avanza una columna para imprimir regla salarial.
                    col += 1
                
                                                
                # Si el departamento es diferente ó es el primer registro.
                if _departamento != query_line['departamento'] or _primer_departamento == True:
                    
                    
                    if  _primer_departamento == True:
                        
                        # Avanza una línea.
                        row += 1
                    
                        # Coloca tipo de planilla.
                        ws.cell(row=row, column=1).value  = query_line['categoria_planilla'] + ' Departamento' or None
                        self.style_range(ws, 'A'+str(row)+':D'+str(row), alignment=al1, font=font)

                        # Avanza una línea.
                        row += 2
                        
                        # Imprime nombre del departamento.
                        ws.cell(row=row, column=1).value  = query_line['departamento'] + ' Departamento' or None
                        self.style_range(ws, 'A'+str(row)+':D'+str(row), alignment=al1, font=font)
                        
                        row += 1
                        
                    elif _tipo_planilla != query_line['categoria_planilla']:
                          
                        _cambia_categoria = True
                        
                    else:
                        
                        # Avanza una línea.
                        row += 1
            
                        # Imprime nombre del departamento.
                        ws.cell(row=row, column=1).value  = query_line['departamento'] + ' Else' or None
                        self.style_range(ws, 'A'+str(row)+':D'+str(row), alignment=al1, font=font)
                        
                        row += 1
     

                    # Si no es el primer registro.
                    if _primer_departamento == False:
    
                                
                        # Si es segundo registro.
                        if _segundo_departamento: 
    
                            # En la columna B pone la cantidad de registros en este departamento.
                            ws.cell(row=row - 2, column=2).value = "=COUNT("'D10:D'+ str(row - 3) + ")"
                            ws.cell(row=row - 2, column=2).number_format = '#,##0'
                            self.style_range(ws, 'C'+str(row-2)+':C'+str(row-2), alignment=al2, font=font)
                            
                            # Acumula fórmula para sumar subtotales por departamento.
                            total_cell[0] = total_cell[0] + "+ B" + str(row-2)
                            
                            # Desde la columna D inicia a totalizar reglas salariales.
                            col_suma_val = 4
                            
                            # Inicia en la columna 0 a colocar fórmula.
                            reg = 1
                                                  
                            # Para cada subtotal de columna. 
                            for col_suma in sub_cell:
                                
                                # Calcula subtotal por departamento.
                                ws.cell(row=row - 2, column=col_suma_val).value = "=SUM("+col_suma+"10:"+col_suma+ str(row - 3) + ")"
                                ws.cell(row=row - 2, column=col_suma_val).number_format = '#,##0.00'
                                self.style_range(ws, col_suma+str(row-2)+':'+col_suma+str(row-2), alignment=al2, font=font)                            
                                                               
                                #Acumula fórmula para el gran total por columna.                               
                                total_cell[reg] = total_cell[reg] + "+" + col_suma + str(row-2)
                                
                                col_suma_val = col_suma_val + 1
                                
                                reg += 1
                                
                            _segundo_departamento = False
                            
                            first_row_sup = row
                
                            
                            
                            
                            
                            
                            
                            
                            
                            
                            

                            
                            
                            
                            
                            
                            
                            
                            
                            
                        # Si no es segundo registro.
                        else:
                            

                            if  _cambia_categoria == True:
                            
                                # Coloca cantidad de registros en este departamento.
                                ws.cell(row=row, column=2).value = "=COUNT("'D' + str(first_row_sup) + ":D"+ str(row - 1) + ")"
                                ws.cell(row=row, column=2).number_format = '#,##0'
                                self.style_range(ws, 'C'+str(row-1)+':C'+str(row-1), alignment=al2, font=font)
                                
                                # Acumula cantidad departamento en cantidad total.
                                total_cell[0] = total_cell[0] + "+ B" + str(row)
                                
                                # Desde la columna D totaliza reglas salariales.
                                col_suma_val = 4
        
                                # Inicia en la columna 0 a colocar fórmula. 
                                reg = 1
                                    
                                # Desde la columna D pone cada subtotal de columna.
                                for col_suma in sub_cell:
                                    
                                    ws.cell(row=row, column=col_suma_val).value = "=SUM("+col_suma + str(first_row_sup) + ":"+col_suma+ str(row - 1) + ")"
                                    ws.cell(row=row, column=col_suma_val).number_format = '#,##0.00'
                                    self.style_range(ws, 'C'+str(row-1)+':'+col_suma+str(row-1), alignment=al2, font=font)
                                    
                                    total_cell[reg] = total_cell[reg] + "+" + col_suma + str(row)
                                    
                                    col_suma_val = col_suma_val + 1
        
                                    # Se mueve de columna.
                                    reg += 1
                                    
                                
                                row += 2
                                
                                
                                ws.cell(row=row, column=1).value = "Subtotal:"
                                 
                                ws.cell(row=row, column=2).value = "="+total_cell[0]
                                ws.cell(row=row, column=2).number_format = '#,##0'
                                    
                                self.style_range(ws, 'C'+str(row)+':'+col_suma+str(row), alignment=al2, font=font)
                                      
                                col_suma_val = 4
                        
                                reg = 1
                                                  
                                # Se mueve en cada columna para colocar valor.
                                for col_suma in sub_cell:

                                    ws.cell(row=row, column=col_suma_val).value = "="+total_cell[reg]
                                    ws.cell(row=row, column=col_suma_val).number_format = '#,##0.00'
                                    self.style_range(ws, 'C'+str(row)+':'+col_suma+str(row+1), alignment=al2, font=font)
                                      
                                    col_suma_val = col_suma_val + 1
                                    
                                    reg += 1
                                    
                                
                                
                                    
                                # Avanza una línea.
                                row += 2
                                
                                ws.cell(row=row, column=1).value  = query_line['categoria_planilla'] + ' Planilla' or None
                                self.style_range(ws, 'A'+str(row)+':D'+str(row), alignment=al1, font=font)

                                # Avanza una línea.
                                row += 1
                                
                                    
                                _cambia_categoria = False
                            
                            else:
                                    
                                
                                # Asigna totales de departamento.
                                # En la columna B pone la cantidad de registros en este grupo.
                                ws.cell(row=row - 2, column=2).value = "=COUNT("'D' + str(first_row_sup) + ":D"+ str(row - 3) + ")"
                                ws.cell(row=row - 2, column=2).number_format = '#,##0'
                                self.style_range(ws, 'C'+str(row-2)+':C'+str(row-2), alignment=al2, font=font)
                                
                                total_cell[0] = total_cell[0] + "+ B" + str(row-2)
                                
                                # Desde la columna D totaliza reglas salariales.
                                col_suma_val = 4
        
                                # Inicia en la columna 0 a colocar fórmula. 
                                reg = 1
                                    
                                # Desde la columna D pone cada subtotal de columna.
                                for col_suma in sub_cell:
                                    
                                    ws.cell(row=row - 2, column=col_suma_val).value = "=SUM("+col_suma + str(first_row_sup) + ":"+col_suma+ str(row - 3) + ")"
                                    ws.cell(row=row - 2, column=col_suma_val).number_format = '#,##0.00'
                                    self.style_range(ws, 'C'+str(row-2)+':'+col_suma+str(row-2), alignment=al2, font=font)
                                    
                                    total_cell[reg] = total_cell[reg] + "+" + col_suma + str(row-2)
                                    
                                    col_suma_val = col_suma_val + 1
        
                                    # Se mueve de columna.
                                    reg += 1      
                    
                            
                            if _tipo_planilla != query_line['categoria_planilla']:


                                # Avanza una línea.
                                row += 1
                    
                                # Imprime nombre del departamento.
                                ws.cell(row=row, column=1).value  = query_line['departamento'] + 'SI' or None
                                self.style_range(ws, 'A'+str(row)+':D'+str(row), alignment=al1, font=font)
                                
                                row += 1
     
    
                            
                            
                            
                            first_row_sup = row
                            
                            
                            
                            
                            

                            
                                
                                                        
                        
                    _primer_departamento = False
                    
                
                
                # Línea de detalle de nómina, 
                # Coloca puesto del empleado.              
                ws.cell(row=row, column=1).value = query_line['puesto'] or None
                self.style_range(ws, ws.cell(row=row, column=1).coordinate+':'+ws.cell(row=row, column=1).coordinate, alignment=al3, font=font2)
                
                # Coloca nombre del empleado. 
                ws.cell(row=row, column=2).value = query_line['empleado'] or None
                self.style_range(ws, ws.cell(row=row, column=2).coordinate+':'+ws.cell(row=row, column=2).coordinate, alignment=al3, font=font2)
                
                # Coloca fecha de ingreso del empleado.
                ws.cell(row=row, column=3).value = query_line['date_start'] or None
                self.style_range(ws, ws.cell(row=row, column=3).coordinate+':'+ws.cell(row=row, column=3).coordinate, alignment=al4, font=font2)
                
                # Coloca valor de la regla salarial.
                ws.cell(row=row, column=col).value  = query_line['total']
                self.style_range(ws, ws.cell(row=row, column=col).coordinate+':'+ws.cell(row=row, column=col).coordinate, alignment=al4, font=font2)
                ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    
    

                # Asigna el empleado.
                _empleado = query_line['empleado']
                
                # Asigna el departamento.
                _departamento = query_line['departamento']
    
                # Asigna el departamento.
                _tipo_planilla = query_line['categoria_planilla']
                
            
        # Si no se debe agrupar por tipo de planilla.
        else:
            

            # Inicia a contar desde el primer registro de detalle.
            first_row_sup = 8
                    
            # Lee cada registro.
            for query_line in self.env.cr.dictfetchall():
    
                # Si el empleado es diferente.
                if _empleado != query_line['empleado']:
                    
                    # En columna 4 inicia a imprimir reglas salariales del empleado.
                    col = 4
                    
                    # Avanza una línea.
                    row += 1
                    
                # Si el empleado es igual.
                else:
                    
                    # Avanza una columna para imprimir regla salarial.
                    col += 1
            
                # Si el departamento es diferente ó es primero.
                if _departamento != query_line['departamento'] or _primer_departamento == True:
                    
                    # Avanza una línea.
                    row += 1
        
                    # Coloca con formato el nombre del departamento.
                    ws.cell(row=row, column=1).value  = query_line['departamento'] or None
                    self.style_range(ws, 'A'+str(row)+':D'+str(row), alignment=al1, font=font)
                    row += 1
     
                    # Si no es el primer registro.
                    if _primer_departamento == False:
                        
                        # Si es segundo registro.
                        if _segundo_departamento: 
    
                            # En la columna B pone la cantidad de registros en este grupo.
                            ws.cell(row=row - 2, column=2).value = "=COUNT("'D8:D'+ str(row - 3) + ")"
                            ws.cell(row=row - 2, column=2).number_format = '#,##0'
                            self.style_range(ws, 'C'+str(row-2)+':C'+str(row-2), alignment=al2, font=font)
                            
                            total_cell[0] = total_cell[0] + "+ B" + str(row-2)
                            
                            # Desde la columna D totaliza reglas salariales.
                            col_suma_val = 4
                            
                            # Inicia en la columna 0 a colocar fórmula.
                            reg = 1
                                                  
                            # Para cada subtotal de columna. 
                            for col_suma in sub_cell:
                                
                                ws.cell(row=row - 2, column=col_suma_val).value = "=SUM("+col_suma+"8:"+col_suma+ str(row - 3) + ")"
                                ws.cell(row=row - 2, column=col_suma_val).number_format = '#,##0.00'
                                self.style_range(ws, col_suma+str(row-2)+':'+col_suma+str(row-2), alignment=al2, font=font)                            

                                total_cell[reg] = total_cell[reg] + "+" + col_suma + str(row-2)
                                
                                col_suma_val = col_suma_val + 1
                                
                                reg += 1
                                
                            _segundo_departamento = False
                            
                            first_row_sup = row
                        
                        # Si no es segundo registro.
                        else:
                        
                            # En la columna B pone la cantidad de registros en este grupo.
                            ws.cell(row=row - 2, column=2).value = "=COUNT("'D' + str(first_row_sup) + ":D"+ str(row - 3) + ")"
                            ws.cell(row=row - 2, column=2).number_format = '#,##0'
                            self.style_range(ws, 'C'+str(row-2)+':C'+str(row-2), alignment=al2, font=font)
                            
                            total_cell[0] = total_cell[0] + "+ B" + str(row-2)
                            
                            # Desde la columna D totaliza reglas salariales.
                            col_suma_val = 4
    
                            # Inicia en la columna 0 a colocar fórmula. 
                            reg = 1
                                
                            # Desde la columna D pone cada subtotal de columna.        
                            for col_suma in sub_cell:
                                
                                ws.cell(row=row - 2, column=col_suma_val).value = "=SUM("+col_suma + str(first_row_sup) + ":"+col_suma+ str(row - 3) + ")"
                                ws.cell(row=row - 2, column=col_suma_val).number_format = '#,##0.00'
                                self.style_range(ws, 'C'+str(row-2)+':'+col_suma+str(row-2), alignment=al2, font=font)
                                
                                total_cell[reg] = total_cell[reg] + "+" + col_suma + str(row-2)
                                
                                col_suma_val = col_suma_val + 1
    
                                # Se mueve de columna.
                                reg += 1
                                
                            first_row_sup = row
                        
                    _primer_departamento = False
                               
                # Línea de detalle de nómina.
                # Coloca puesto del empleado.              
                ws.cell(row=row, column=1).value = query_line['puesto'] or None
                self.style_range(ws, ws.cell(row=row, column=1).coordinate+':'+ws.cell(row=row, column=1).coordinate, alignment=al3, font=font2)
                
                # Coloca nombre del empleado. 
                ws.cell(row=row, column=2).value = query_line['empleado'] or None
                self.style_range(ws, ws.cell(row=row, column=2).coordinate+':'+ws.cell(row=row, column=2).coordinate, alignment=al3, font=font2)
                
                # Coloca fecha de ingreso del empleado.
                ws.cell(row=row, column=3).value = query_line['date_start'] or None
                self.style_range(ws, ws.cell(row=row, column=3).coordinate+':'+ws.cell(row=row, column=3).coordinate, alignment=al4, font=font2)
                
                # Coloca valor de la regla salarial.
                ws.cell(row=row, column=col).value  = query_line['total']
                self.style_range(ws, ws.cell(row=row, column=col).coordinate+':'+ws.cell(row=row, column=col).coordinate, alignment=al4, font=font2)
                ws.cell(row=row, column=col).number_format = '#,##0.00'
    
                # Asigna el empleado.
                _empleado = query_line['empleado']
                
                # Asigna el departamento.
                _departamento = query_line['departamento']
    
                # Asigna el departamento.
                _tipo_planilla = query_line['categoria_planilla']
        
                
        # Último total por departamento.                
        # En la columna B pone la cantidad de registros en este grupo.
        ws.cell(row=row + 1, column=2).value = "=COUNT("'D' + str(first_row_sup) + ":D"+ str(row) + ")"
        ws.cell(row=row + 1, column=2).number_format = '#,##0'
        self.style_range(ws, 'C'+str(row+1)+':C'+str(row+1), alignment=al2, font=font)
        
        # Calcula el total.
        total_cell[0] = total_cell[0] + "+ B" + str(row+1)
        
        ws.cell(row=row + 3, column=1).value = "Gran Total:"
         
        ws.cell(row=row + 3, column=2).value = "="+total_cell[0]
        ws.cell(row=row + 3, column=2).number_format = '#,##0'
            
        self.style_range(ws, 'C'+str(row+3)+':'+col_suma+str(row+3), alignment=al2, font=font)
              
        col_suma_val = 4

        reg = 1
        
        # Se mueve en cada columna para colocar valor.
        for col_suma in sub_cell:
        	
            ws.cell(row=row + 1, column=col_suma_val).value = "=SUM("+col_suma + str(first_row_sup) + ":"+col_suma + str(row) + ")"
            ws.cell(row=row + 1, column=col_suma_val).number_format = '#,##0.00'
            self.style_range(ws, 'C'+str(row+1)+':'+col_suma+str(row+1), alignment=al2, font=font)
            
            total_cell[reg] = total_cell[reg] + "+" + col_suma + str(row+1)
            
            ws.cell(row=row + 3, column=col_suma_val).value = "="+total_cell[reg]
            ws.cell(row=row + 3, column=col_suma_val).number_format = '#,##0.00'
            self.style_range(ws, 'C'+str(row+3)+':'+col_suma+str(row+3), alignment=al2, font=font)
              
            col_suma_val = col_suma_val + 1
            
            reg += 1
            
        celda = ws.cell(row=1, column=col_suma_val-1).coordinate        
        ws.merge_cells('A1:'+celda)
        
        celda = ws.cell(row=2, column=col_suma_val-1).coordinate
        ws.merge_cells('A2:'+celda)
        
        
        
        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)
        
        self.write({
            'state': 'get',
            'name': "rep_planilla_spreadsheet.xlsx",
            'data': out
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'rep_planilla.generate_planilla',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this,
            'views': [(False, 'form')],
            'target': 'new'
        }

    def style_range(self, ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
        """
        Apply styles to a range of cells as if they were a single cell.
        :param ws:  Excel worksheet instance
        :param range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        :param fill: An openpyxl PatternFill or GradientFill
        :param font: An openpyxl Font object
        """

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        rows = list(ws[cell_range])
        for cell in rows[0]:
            cell.border = top
        for cell in rows[-1]:
            cell.border = bottom
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border =  left
            r.border =  right
            if fill:
                for c in row:
                    c.fill = fill
            if font:
                for c in row:
                    c.font = font

            if alignment:
                for c in row:
                    c.alignment = alignment